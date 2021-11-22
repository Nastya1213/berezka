import sys
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5 import QtWidgets
from unicodedata import normalize
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook
from random import shuffle


# Функция для извлечения данных из столбца и записи в список
# 1-й аргумент - это буквенный номер столбца
# 2-й аргумент - это в какой список нужно поместить
# 3-й и 4-й аргументы - это с какой по счету строки начинаются данные и какой заканчиваются
def answr(column, what_list, start, finish):
    for i in range(start, finish):  # цикл чтобы пройти от начала и до конца
        a = sheet[f'{column}{i}'].value  # Sheet у нас это лист, из него мы получаем значение в столбце,
        # который мы указали в аргкменте, по номеру i
        what_list.append(normalize("NFKD", a.lower()))  # в таблице, имелись неразрывные пробелы,
        # которые в списке выглядили как \xa0


# Извлекаем и таблицы Excel нужные данные для теста "60 частых ошибок"
wb = load_workbook('words.xlsx')
sheet = wb['Лист1']
right_answers_60 = []  # правильные ответы на тест "60 частых ошибок"
wrong_answers_60 = []  # неправильные ответы на тест "60 частых ошибок"
answr('B', right_answers_60, 2, 62)
answr('C', wrong_answers_60, 2, 62)
# Извлекаем и таблицы Excel нужные данные для теста "Орфоэпия"
tab = load_workbook('Orthoepy.xlsx')
sheet = tab['Лист1']
right_answers_ort = [] # правильные ответы для теста "Орфоэпия"
wrong_answers_ort = [] # неправильные ответы для теста "Орфоэпия"
answr('A', right_answers_ort, 1, 16)
answr('B', wrong_answers_ort, 1, 16)
# Извлекаем и таблицы Excel нужные данные для теста "Правописание н/нн"
wb = load_workbook('n_nn.xlsx')
sheet = wb['Лист1']
right_answers_doubled_n = []
wrong_answers_doubled_n = []
answr('A', right_answers_doubled_n, 1, 11)
answr('B', wrong_answers_doubled_n, 1, 11)
# создаем новый excel-файл для записи результатов
res_wrkb = openpyxl.Workbook()
# добавляем новый лист
res_wrkb.create_sheet(title='Первый лист', index=0)
# получаем лист, с которым будем работать
sht = res_wrkb['Первый лист']
sht['A1'] = 'Название теста'
sht['B1'] = 'Выполнено правильно %'
sht['C1'] = 'Сколько баллов'
sht['A2'] = '60 частых ошибок'
sht['A3'] = 'Тест на орфоэпию'
sht['A4'] = 'Правописание н/нн'


class Doubledn(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        # загружаем дизайн
        uic.loadUi('dialog.ui', self)
        self.setWindowTitle('Правописание н/нн')
        self.right_count = 0 # количество правильных ответов
        self.sheet = sht # excel-лист, на который мы записываем результат пользователя
        self.users_answers = []  # список для ответов пользователя
        self.temporary_lst = []  # времменный список, чтобы в users_answers вносить не каждый ответ,
        # который,еще думая, выбирал пользователь, а последний его ответ перед нажатием
        # на кнопку (pushButton) 'следующий вопрос'
        self.count = 0  # номер вопроса
        # Инициализация списков
        self.right_answers = right_answers_doubled_n
        self.wrong_answers = wrong_answers_doubled_n
        # добавляем в список с вариантами 1 правильный и 1 неправильный ответы
        self.options = [self.right_answers[0], self.wrong_answers[0]]
        # перемешиваем варианты
        shuffle(self.options)
        # меняем текст на radioButton на наши варианты ответов
        self.rb1.setText(self.options[0])
        self.rb2.setText(self.options[1])
        self.rb3.setText('Пропустить')
        self.count += 1
        # очищаем список вариантов
        self.options.clear()
        self.pushButton.clicked.connect(self.run)
        self.rb1.toggled.connect(self.answer_user)
        self.rb2.toggled.connect(self.answer_user)
        self.rb3.toggled.connect(self.answer_user)
        # по умолчанию мы выбираем кнопку пропустить,
        # чтобы при быстром пролистывании никакие ответы не были записаны
        self.rb3.setChecked(True)

    # функция собирает все ответы пользователя, если пользователь сначало выбрал один ответ,
    # но потом передумал, она все равно запишет его во временный список, из которого в функции run
    # мы добавим только окончательный ответ в users_answers
    def answer_user(self):
        self.temporary_lst.append(self.sender().text())

    def run(self):
        # если вопросы закончились
        if self.count == 10:
            self.users_answers.append(self.temporary_lst[-1])
            # пробегаем по индексам ответов пользователя
            for i in range(10):
                # если ответ пользователя такой же как и в правильных ответах,
                # то прибавляем 1 к счетчику
                if self.users_answers[i] == self.right_answers[i]:
                    self.right_count += 1
            # записываем количество правильных ответов
            self.sheet['C4'] = self.right_count
            # Правильных ответов в процентах
            # 1 правильный ответ = 10%
            self.sheet['B4'] = self.right_count * 10
            res_wrkb.save('Result.xlsx')
            # закрываем диалоговое окно
            self.close()
        # если вопросы еще не закончились
        if self.count < 10:
            self.options.append(self.right_answers[self.count])
            self.options.append(self.wrong_answers[self.count])
            shuffle(self.options)
            self.rb1.setText(self.options[0])
            self.rb2.setText(self.options[1])
            self.count += 1
            self.options.clear()
            # добавляе в спиосок ответов пользователя, его последний ответ
            self.users_answers.append(self.temporary_lst[-1])
            # По умолчанию "пропустить"
            self.rb3.setChecked(True)


class Orthoepy(QtWidgets.QDialog):
    def __init__(self):
        super(Orthoepy, self).__init__()
        # загружаем дизайн
        uic.loadUi('dialog.ui', self)
        self.setWindowTitle('Орфоэпия')
        self.users_answers = []  # список для ответов пользователя
        self.temporary_lst = []  # времменный список, чтобы в users_answers вносить не каждый ответ,
        # который,еще думая, выбирал пользователь, а последний его ответ перед нажатием
        # на кнопку (pushButton) 'следующий вопрос'
        self.count = 0  # кол-во пройденных вопросов
        self.right_count = 0 # кол-во правильных ответов пользователя
        self.right_answers = right_answers_ort
        self.wrong_answers = wrong_answers_ort
        self.sheet = sht  # excel-лист, на который мы записываем результат
        self.options = [self.right_answers[0], self.wrong_answers[0]]
        shuffle(self.options)
        self.rb1.setText(self.options[0])
        self.rb2.setText(self.options[1])
        self.rb3.setText('Пропустить')
        self.count += 1
        self.options.clear()
        self.pushButton.clicked.connect(self.run)
        self.rb1.toggled.connect(self.answer_user)
        self.rb2.toggled.connect(self.answer_user)
        self.rb3.toggled.connect(self.answer_user)
        self.rb3.setChecked(True)

    def answer_user(self):
        self.temporary_lst.append(self.sender().text())

    def run(self):
        # если вопросы закончились
        if self.count == 15:
            # закрываем диалоговое окно
            self.close
            self.count += 1
            # проходимся по индексам и сравниваем их
            for i in range(14):
                # если ответ пользователя соответствует правильному
                if self.users_answers[i] == self.right_answers[i]:
                    # к количеству правильных ответов прибавляем 1
                    self.right_count += 1
            # в ячейку добавляем сколько правильных ответов в %
            # 3 правильных ответов = 20%
            self.sheet['B3'] = round(self.right_count / 3 * 20)
            # в ячейку добавляем кол-во правильных ответов пользователя
            self.sheet['C3'] = self.right_count
            res_wrkb.save('Result.xlsx')

        # вопросы остались
        if self.count < 15:
            self.options.append(self.right_answers[self.count])
            self.options.append(self.wrong_answers[self.count])
            shuffle(self.options)
            self.rb1.setText(self.options[0])
            self.rb2.setText(self.options[1])
            self.count += 1
            self.options.clear()
            self.users_answers.append(self.temporary_lst[-1])
            self.rb3.setChecked(True)


class Test60(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        self.right_answers = right_answers_60
        self.wrong_answers = wrong_answers_60
        self.sheet = sht  # excel-лист, на который мы записываем результат
        self.options = []
        self.count = 0
        self.right_count = 0
        self.users_answers = []
        self.temporary_lst = []
        uic.loadUi('dialog.ui', self)
        self.options.append(self.right_answers[self.count])
        self.options.append(self.wrong_answers[self.count])
        shuffle(self.options)
        self.rb1.setText(self.options[0])
        self.rb2.setText(self.options[1])
        self.count += 1
        self.options.clear()
        self.pushButton.clicked.connect(self.run)
        self.rb1.toggled.connect(self.answer_user)
        self.rb2.toggled.connect(self.answer_user)
        self.rb3.toggled.connect(self.answer_user)
        self.rb3.setChecked(True)

    def answer_user(self):
        self.temporary_lst.append(self.sender().text())

    def run(self):
        if self.count == 58:
            self.pushButton.setText('Завершить и узнать результат!')
            self.options.append(self.right_answers[self.count])
            self.options.append(self.wrong_answers[self.count])
            shuffle(self.options)
            self.rb1.setText(self.options[0])
            self.rb2.setText(self.options[1])
            self.options.clear()
            self.count += 1
            self.users_answers.append(self.temporary_lst[-1])
        # если все все вопросы решены
        if self.count == 60:
            # пробегаем по ответам пользователя
            for i in range(59):
                # если ответ пользователя соответсвует правильнуму ответу,
                # то прибавляем 1 к счетчику
                if self.users_answers[i] == self.right_answers[i]:
                    self.right_count += 1
            # записываем в ячейку % правильных ответов
            # 3 правильных ответа = 5%
            self.sheet['B2'] = round(self.right_count / 3 * 5)
            # записываем в ячейку количество правильных ответов
            self.sheet['C2'] = self.right_count
            res_wrkb.save('Result.xlsx')
            # закрываем диалоговое окно
            self.close()
        # остались ещё вопросы
        else:
            self.options.append(self.right_answers[self.count])
            self.options.append(self.wrong_answers[self.count])
            shuffle(self.options)
            self.rb1.setText(self.options[0])
            self.rb2.setText(self.options[1])
            self.options.clear()
            self.count += 1
            self.users_answers.append(self.temporary_lst[-1])
            self.rb3.setChecked(True)


class Names(QtWidgets.QDialog):
    def __init__(self):
        self.sheet = sht
        super().__init__()
        uic.loadUi('users_name.ui', self)
        self.pushButton.clicked.connect(self.run)

    def run(self):
        self.sheet['F1'] = self.input_name.text()
        res_wrkb.save('Result.xlsx')
        self.close()


class MyWidget(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('design.ui', self)  # Загружаем дизайн
        self.test_60.clicked.connect(self.run_test_60)
        self.test_orthoepy.clicked.connect(self.run_test_orthoepy)
        self.test_doubled_n.clicked.connect(self.run_test_doubled_n)

    def run_test_60(self):
        test = Test60()
        test.exec_()

    def run_test_orthoepy(self):
        test = Orthoepy()
        test.exec_()

    def run_test_doubled_n(self):
        test = Doubledn()
        test.exec_()


def except_hook(cls, exception, traceback):
    """Функция для отслеживания ошибок PyQt5"""
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWidget()
    ex.show()
    n = Names()
    n.show()
    sys.excepthook = except_hook
    sys.exit(app.exec_())